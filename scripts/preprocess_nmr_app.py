#!/usr/bin/env python3
"""
Preprocess the Hiller lab NMR-APP dataset for the robustness-dynamics pipeline.

Reads the xlsx file with per-residue R1, R2, hetNOE for 9 designed proteins
from the NMR-APP pipeline (Muntener et al. 2026).

Since these are de novo designed proteins with no PDB structures, we need
AlphaFold2 (or ESMFold) predicted structures. This script:
  1. Parses the xlsx (sequences + per-residue relaxation)
  2. Writes FASTA files for AF2 prediction
  3. Creates per-protein directory structure for the pipeline
  4. Optionally folds sequences with ESMFold (if --fold flag given)

After running this script:
  - If AF2 PDBs are provided via --pdb_dir, they are symlinked in
  - If --fold is given, ESMFold is used to generate structures
  - Otherwise, FASTA files are written and you must run AF2 externally

Output directory structure (per protein):
  {output_dir}/proteins/{protein_id}/
    {protein_id}.pdb              # AF2/ESMFold structure
    {protein_id}.fasta            # sequence
    {protein_id}_Bfactor.tsv      # 1 - hetNOE (primary target, matches pipeline convention)
    {protein_id}_hetNOE.tsv       # raw hetNOE values
    {protein_id}_R1.tsv           # R1 values
    {protein_id}_R2.tsv           # R2 values
    {protein_id}_R2R1.tsv         # R2/R1 ratio
    {protein_id}_pLDDT.tsv        # pLDDT from predicted structure
    .done

Usage:
  # Step 1: parse xlsx, write FASTA + relaxation TSVs
  python preprocess_nmr_app.py \\
      --xlsx /path/to/Relaxation_Data_9_proteins_NMR_APP.xlsx \\
      --output_dir /path/to/nmr_app_processed

  # Step 2 (option A): fold with ESMFold in-process (needs GPU + esm package)
  python preprocess_nmr_app.py \\
      --xlsx /path/to/Relaxation_Data_9_proteins_NMR_APP.xlsx \\
      --output_dir /path/to/nmr_app_processed \\
      --fold

  # Step 2 (option B): provide pre-computed AF2 PDBs
  python preprocess_nmr_app.py \\
      --xlsx /path/to/Relaxation_Data_9_proteins_NMR_APP.xlsx \\
      --output_dir /path/to/nmr_app_processed \\
      --pdb_dir /path/to/af2_predictions

Ref: Muntener et al. 2026, bioRxiv 10.64898/2026.02.16.706194
"""

import argparse
import sys
import numpy as np
import pandas as pd
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)


def parse_xlsx(xlsx_path: str) -> dict:
    """Parse the NMR-APP xlsx file.

    Returns dict: protein_id -> {
        'sequence': str,
        'relaxation': DataFrame with columns [position, R1, R1err, R2, R2err, hetNOE, hetNOEerr]
    }
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # Parse sequences from the Sequences sheet
    ws_seq = wb["Sequences"]
    sequences = {}
    for row in ws_seq.iter_rows(values_only=True):
        if row[0] is not None and str(row[0]).startswith("p"):
            pid = str(row[0]).strip()
            # Sequence may be split across columns
            seq_parts = [str(v) for v in row[1:] if v is not None and str(v).strip()]
            seq = "".join(seq_parts).replace(" ", "").replace("\n", "")
            if len(seq) > 10:  # sanity check
                sequences[pid] = seq

    # Parse relaxation data from per-protein sheets
    proteins = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "Sequences":
            continue

        pid = sheet_name.strip()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Header row
        header = [str(h).strip() if h is not None else "" for h in rows[0]]

        # Data rows
        data_rows = []
        for row in rows[1:]:
            if row[0] is None:
                continue
            vals = {}
            for j, col_name in enumerate(header):
                if j < len(row) and row[j] is not None:
                    if col_name.lower() in ("residue", ""):
                        vals["position"] = int(row[j])
                    else:
                        try:
                            vals[col_name] = float(row[j])
                        except (ValueError, TypeError):
                            vals[col_name] = np.nan
            if "position" in vals:
                data_rows.append(vals)

        df = pd.DataFrame(data_rows)

        if pid not in sequences:
            print(f"  WARNING: no sequence found for {pid}, skipping")
            continue

        proteins[pid] = {
            "sequence": sequences[pid],
            "relaxation": df,
        }

    wb.close()
    return proteins


def extract_plddt_from_pdb(pdb_path: str) -> list:
    """Extract per-residue pLDDT from AF2/ESMFold PDB B-factor column."""
    plddt_vals = []
    seen = set()
    with open(pdb_path) as f:
        for line in f:
            if line.startswith("ATOM") and line[12:16].strip() == "CA":
                resnum = int(line[22:26].strip())
                if resnum not in seen:
                    bfac = float(line[60:66].strip())
                    plddt_vals.append((resnum, bfac))
                    seen.add(resnum)
    return plddt_vals


def fold_with_esmfold(sequence: str, output_pdb: str):
    """Fold a sequence with ESMFold and save as PDB."""
    import torch
    import esm

    # Lazy-load model (cached after first call)
    if not hasattr(fold_with_esmfold, "_model"):
        print("  Loading ESMFold model (first call, may take a minute)...")
        fold_with_esmfold._model = esm.pretrained.esmfold_v1()
        fold_with_esmfold._model = fold_with_esmfold._model.eval()
        if torch.cuda.is_available():
            fold_with_esmfold._model = fold_with_esmfold._model.cuda()

    model = fold_with_esmfold._model

    with torch.no_grad():
        output = model.infer_pdb(sequence)

    with open(output_pdb, "w") as f:
        f.write(output)


def create_protein_dir(pid: str, data: dict, output_dir: Path,
                       pdb_dir: Path = None, do_fold: bool = False):
    """Create the per-protein directory with all required files."""
    prot_dir = output_dir / "proteins" / pid
    prot_dir.mkdir(parents=True, exist_ok=True)

    seq = data["sequence"]
    df = data["relaxation"]

    # Write FASTA
    fasta_path = prot_dir / f"{pid}.fasta"
    with open(fasta_path, "w") as f:
        f.write(f">{pid}\n{seq}\n")

    # Handle PDB structure
    pdb_path = prot_dir / f"{pid}.pdb"
    if pdb_dir is not None:
        # Look for pre-computed PDB
        candidates = [
            pdb_dir / f"{pid}.pdb",
            pdb_dir / f"{pid}_relaxed.pdb",
            pdb_dir / f"{pid}_unrelaxed.pdb",
        ]
        found = False
        for cand in candidates:
            if cand.exists():
                if pdb_path.is_symlink() or pdb_path.exists():
                    pdb_path.unlink()
                pdb_path.symlink_to(cand.resolve())
                found = True
                break
        if not found:
            print(f"  WARNING: No PDB found for {pid} in {pdb_dir}")
    elif do_fold:
        if not pdb_path.exists():
            print(f"  Folding {pid} with ESMFold...")
            fold_with_esmfold(seq, str(pdb_path))
    else:
        if not pdb_path.exists():
            print(f"  No PDB for {pid} — run with --fold or provide --pdb_dir")

    # Write relaxation TSVs
    # hetNOE as primary target (1 - hetNOE = flexibility, like 1 - S2)
    if "hetNOE" in df.columns:
        bfactor_df = pd.DataFrame({
            "position": df["position"],
            "bfactor": 1.0 - df["hetNOE"],  # high = flexible
        }).dropna(subset=["bfactor"])
        bfactor_df.to_csv(prot_dir / f"{pid}_Bfactor.tsv",
                          sep="\t", index=False)

        # Also write raw hetNOE
        het_df = pd.DataFrame({
            "position": df["position"],
            "hetNOE": df["hetNOE"],
        })
        if "hetNOEerr" in df.columns:
            het_df["hetNOE_err"] = df["hetNOEerr"]
        het_df.dropna(subset=["hetNOE"]).to_csv(
            prot_dir / f"{pid}_hetNOE.tsv", sep="\t", index=False)

    # R1
    if "R1" in df.columns:
        r1_df = pd.DataFrame({"position": df["position"], "R1": df["R1"]})
        if "R1err" in df.columns:
            r1_df["R1_err"] = df["R1err"]
        r1_df.dropna(subset=["R1"]).to_csv(
            prot_dir / f"{pid}_R1.tsv", sep="\t", index=False)

    # R2
    if "R2" in df.columns:
        r2_df = pd.DataFrame({"position": df["position"], "R2": df["R2"]})
        if "R2err" in df.columns:
            r2_df["R2_err"] = df["R2err"]
        r2_df.dropna(subset=["R2"]).to_csv(
            prot_dir / f"{pid}_R2.tsv", sep="\t", index=False)

    # R2/R1 ratio
    if "R1" in df.columns and "R2" in df.columns:
        mask = (df["R1"] > 0) & df["R1"].notna() & df["R2"].notna()
        r2r1_df = pd.DataFrame({
            "position": df.loc[mask, "position"],
            "R2R1": df.loc[mask, "R2"] / df.loc[mask, "R1"],
        })
        r2r1_df.to_csv(prot_dir / f"{pid}_R2R1.tsv", sep="\t", index=False)

    # pLDDT from PDB (if available)
    if pdb_path.exists():
        plddt_vals = extract_plddt_from_pdb(str(pdb_path))
        if plddt_vals:
            plddt_df = pd.DataFrame(plddt_vals, columns=["position", "plddt"])
            plddt_df.to_csv(prot_dir / f"{pid}_pLDDT.tsv",
                            sep="\t", index=False)

    # Sentinel
    (prot_dir / ".done").touch()

    n_res = len(df)
    has_pdb = pdb_path.exists()
    print(f"  {pid}: {len(seq)} residues, {n_res} with relaxation data, "
          f"PDB={'yes' if has_pdb else 'NO'}")


def main():
    parser = argparse.ArgumentParser(
        description="Preprocess NMR-APP designed protein relaxation data")
    parser.add_argument("--xlsx", required=True,
                        help="Path to Relaxation_Data_9_proteins_NMR_APP.xlsx")
    parser.add_argument("--output_dir", required=True,
                        help="Output directory")
    parser.add_argument("--pdb_dir", default=None,
                        help="Directory with pre-computed AF2/ESMFold PDBs")
    parser.add_argument("--fold", action="store_true",
                        help="Fold sequences with ESMFold (needs GPU + esm)")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    pdb_dir = Path(args.pdb_dir) if args.pdb_dir else None

    print(f"=== NMR-APP Preprocessing ===")
    print(f"Input:  {args.xlsx}")
    print(f"Output: {output_dir}")

    # Parse xlsx
    proteins = parse_xlsx(args.xlsx)
    print(f"\nParsed {len(proteins)} proteins")

    # Create directories
    print("\nCreating protein directories:")
    for pid, data in sorted(proteins.items()):
        create_protein_dir(pid, data, output_dir, pdb_dir, args.fold)

    # Write summary
    summary = []
    for pid, data in sorted(proteins.items()):
        df = data["relaxation"]
        summary.append({
            "protein_id": pid,
            "seq_length": len(data["sequence"]),
            "n_residues_with_data": len(df),
            "mean_R1": df["R1"].mean() if "R1" in df.columns else np.nan,
            "mean_R2": df["R2"].mean() if "R2" in df.columns else np.nan,
            "mean_hetNOE": df["hetNOE"].mean() if "hetNOE" in df.columns else np.nan,
        })
    summary_df = pd.DataFrame(summary)
    summary_df.to_csv(output_dir / "nmr_app_summary.tsv", sep="\t", index=False)
    print(f"\nSummary written to {output_dir / 'nmr_app_summary.tsv'}")

    # Write combined FASTA for batch AF2 prediction
    fasta_all = output_dir / "all_sequences.fasta"
    with open(fasta_all, "w") as f:
        for pid, data in sorted(proteins.items()):
            f.write(f">{pid}\n{data['sequence']}\n")
    print(f"Combined FASTA: {fasta_all}")

    print(f"\nDone. Next steps:")
    if pdb_dir or args.fold:
        print("  Structures available. Run robustness computation:")
        print(f"  python compute_robustness.py --scorer thermompnn \\")
        print(f"      --atlas_dir {output_dir} --output_dir <robustness_dir> \\")
        print(f"      --batch --device cuda --thermompnn_dir $THERMOMPNN_DIR")
    else:
        print(f"  1. Generate AF2 structures from {fasta_all}")
        print(f"  2. Re-run with --pdb_dir <af2_output> to link structures")
        print(f"  3. Then run robustness computation")


if __name__ == "__main__":
    main()
